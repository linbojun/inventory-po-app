from typing import List, Dict, Tuple, Optional
from openpyxl import load_workbook
from PyPDF2 import PdfReader
from decimal import Decimal, InvalidOperation
import re

from app.schemas import ProductCreate

def parse_excel(file_path: str) -> Tuple[List[ProductCreate], List[Dict]]:
    """
    Parse Excel file and extract product data.
    Returns tuple of (products, errors)
    """
    products = []
    errors = []
    
    try:
        workbook = load_workbook(file_path, data_only=True)
        sheet = workbook.active
        
        # Find header row
        header_row = None
        header_map = {}
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=False), start=1):
            row_values = [cell.value for cell in row if cell.value]
            if not row_values:
                continue
                
            # Check if this looks like a header row
            row_lower = [str(val).lower().strip() if val else "" for val in row_values]
            if any(keyword in row_lower for keyword in ['product_id', 'product id', 'id', 'product']):
                header_row = row_idx
                # Map column names (case-insensitive, flexible)
                for col_idx, cell_value in enumerate(row_values, start=1):
                    if cell_value:
                        cell_lower = str(cell_value).lower().strip()
                        if 'product_id' in cell_lower or ('product' in cell_lower and 'id' in cell_lower):
                            header_map['product_id'] = col_idx
                        elif 'name' in cell_lower:
                            header_map['name'] = col_idx
                        elif 'brand' in cell_lower:
                            header_map['brand'] = col_idx
                        elif 'price' in cell_lower:
                            header_map['price'] = col_idx
                        elif 'stock' in cell_lower:
                            header_map['stock'] = col_idx
                        elif 'order' in cell_lower and 'qty' in cell_lower:
                            header_map['order_qty'] = col_idx
                        elif 'image' in cell_lower or 'url' in cell_lower:
                            header_map['image_url'] = col_idx
                        elif 'remark' in cell_lower or 'note' in cell_lower:
                            header_map['remarks'] = col_idx
                break
        
        if not header_row:
            errors.append({"row": 0, "error": "Could not find header row in Excel file"})
            return products, errors
        
        # Parse data rows
        for row_idx, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            row_values = list(row)
            if not any(row_values):  # Skip empty rows
                continue
            
            try:
                product_id = str(row_values[header_map.get('product_id', 1) - 1]).strip() if header_map.get('product_id') and len(row_values) >= header_map['product_id'] else None
                name = str(row_values[header_map.get('name', 2) - 1]).strip() if header_map.get('name') and len(row_values) >= header_map['name'] else None
                
                if not product_id or not name:
                    errors.append({"row": row_idx, "error": "Missing required fields: product_id and name"})
                    continue
                
                # Extract optional fields
                brand = None
                if header_map.get('brand') and len(row_values) >= header_map['brand']:
                    brand_val = row_values[header_map['brand'] - 1]
                    brand = str(brand_val).strip() if brand_val else None
                
                price = Decimal('0')
                if header_map.get('price') and len(row_values) >= header_map['price']:
                    price_val = row_values[header_map['price'] - 1]
                    try:
                        price = Decimal(str(price_val)) if price_val is not None else Decimal('0')
                    except:
                        price = Decimal('0')
                
                stock = 0
                if header_map.get('stock') and len(row_values) >= header_map['stock']:
                    stock_val = row_values[header_map['stock'] - 1]
                    try:
                        stock = int(float(stock_val)) if stock_val is not None else 0
                    except:
                        stock = 0
                
                order_qty = 0
                if header_map.get('order_qty') and len(row_values) >= header_map['order_qty']:
                    order_qty_val = row_values[header_map['order_qty'] - 1]
                    try:
                        order_qty = int(float(order_qty_val)) if order_qty_val is not None else 0
                    except:
                        order_qty = 0
                
                image_url = None
                if header_map.get('image_url') and len(row_values) >= header_map['image_url']:
                    image_val = row_values[header_map['image_url'] - 1]
                    image_url = str(image_val).strip() if image_val else None
                
                remarks = None
                if header_map.get('remarks') and len(row_values) >= header_map['remarks']:
                    remarks_val = row_values[header_map['remarks'] - 1]
                    remarks = str(remarks_val).strip() if remarks_val else None
                
                product = ProductCreate(
                    product_id=product_id,
                    name=name,
                    brand=brand,
                    price=price,
                    stock=stock,
                    order_qty=order_qty,
                    image_url=image_url,
                    remarks=remarks
                )
                products.append(product)
                
            except Exception as e:
                errors.append({"row": row_idx, "error": f"Error parsing row: {str(e)}"})
        
    except Exception as e:
        errors.append({"row": 0, "error": f"Error reading Excel file: {str(e)}"})
    
    return products, errors


PDF_HEADER_KEYWORDS = {
    "product_id": ["item code", "item#", "item no", "product code", "code"],
    "description": ["description", "item description"],
    "price_each": ["price each", "unit price", "price/ea", "price ea", "price ea.", "price ea"],
    "amount": ["amount", "ext price", "ext.", "extension", "total"],
    "case_pack": ["case pack", "pack", "pkg", "case"],
    "qty": ["qty", "order qty", "quantity", "ord qty"],
}

PDF_FOOTER_MARKERS = [
    "subtotal",
    "tax",
    "total",
    "balance",
    "thank you",
    "payment",
    "invoice total",
]

PDF_CURRENCY_PATTERN = re.compile(r"(-?\$?\s*\(?\d[\d,]*\.\d{2}\)?)")


def parse_pdf(file_path: str) -> Tuple[List[ProductCreate], List[Dict]]:
    """
    Parse PDF file generated from Chinatown Supermarket invoices and extract product rows.
    Returns tuple of (products, errors).
    """
    errors: List[Dict] = []
    product_rows: List[Dict] = []
    products_by_id: Dict[str, Dict] = {}
    
    try:
        reader = PdfReader(file_path)
    except Exception as exc:
        return [], [{"row": 0, "error": f"Error opening PDF file: {exc}"}]
    
    lines_with_meta: List[Tuple[str, int]] = []
    for page_idx, page in enumerate(reader.pages, start=1):
        try:
            page_text = page.extract_text() or ""
        except Exception as exc:
            errors.append({"row": page_idx, "error": f"Could not extract text from page {page_idx}: {exc}"})
            continue
        
        for raw_line in page_text.splitlines():
            lines_with_meta.append((raw_line.rstrip(), page_idx))
    
    if not lines_with_meta:
        errors.append({"row": 0, "error": "PDF file appears to be empty or contains only scanned images."})
        return [], errors
    
    # First try to parse PDFs that match the Chinatown Supermarket layout
    special_parse = _try_parse_chinatown_invoice(lines_with_meta)
    if special_parse is not None:
        return special_parse

    header_index, column_positions = _locate_pdf_header(lines_with_meta)
    if header_index == -1:
        errors.append({
            "row": 0,
            "error": "Could not locate header row with 'Item Code', 'Description', and 'Price Each'. Falling back to pattern-based parsing."
        })
    
    last_product: Optional[Dict] = None
    
    idx = 0
    while idx < len(lines_with_meta):
        raw_line, page_num = lines_with_meta[idx]
        stripped_line = raw_line.strip()
        if not stripped_line:
            idx += 1
            continue
        
        normalized = stripped_line.lower()
        
        # Detect new header blocks (useful for multipage invoices)
        if _looks_like_header(stripped_line):
            header_index = idx
            column_positions = _build_column_positions(stripped_line)
            idx += 1
            continue
        
        if any(marker in normalized for marker in PDF_FOOTER_MARKERS):
            idx += 1
            continue
        
        if header_index != -1 and idx <= header_index:
            idx += 1
            continue
        
        parsed = _parse_line_by_columns(raw_line, column_positions)
        if not parsed:
            parsed = _parse_line_by_pattern(raw_line)
        
        if not parsed:
            if last_product and stripped_line and not any(marker in normalized for marker in PDF_FOOTER_MARKERS):
                # Treat as description continuation
                last_product["name"] = f"{last_product['name']} {stripped_line}".strip()
            else:
                errors.append({
                    "row": f"page {page_num}, line {idx + 1}",
                    "error": f"Unable to parse row: {raw_line}"
                })
            idx += 1
            continue
        
        product_id, name, price = parsed
        if not product_id or not name or price is None:
            errors.append({
                "row": f"page {page_num}, line {idx + 1}",
                "error": f"Incomplete row detected: {raw_line}"
            })
            idx += 1
            continue
        
        record = {
            "product_id": product_id,
            "name": name,
            "price": price,
            "source": f"page {page_num}, line {idx + 1}",
        }
        
        if product_id in products_by_id:
            errors.append({
                "row": record["source"],
                "error": f"Duplicate item code '{product_id}' detected. Using the last occurrence."
            })
            products_by_id[product_id].update(record)
            last_product = products_by_id[product_id]
        else:
            products_by_id[product_id] = record
            product_rows.append(record)
            last_product = record
        
        idx += 1
    
    if not product_rows:
        errors.append({
            "row": 0,
            "error": "Could not extract product data from PDF. Ensure the PDF contains an Item Code/Description/Price Each table."
        })
        return [], errors
    
    products = [
        ProductCreate(
            product_id=row["product_id"],
            name=row["name"],
            brand=None,
            price=row["price"],
            stock=0,
            order_qty=0,
            image_url=None,
            remarks=None,
        )
        for row in product_rows
    ]
    
    return products, errors


def _try_parse_chinatown_invoice(
    lines_with_meta: List[Tuple[str, int]]
) -> Optional[Tuple[List[ProductCreate], List[Dict]]]:
    """
    Best-effort parser for the Chinatown Supermarket invoice layout supplied by the user.
    The layout features multi-line descriptions followed by a pack/quantity summary line
    that contains the price, amount, item code, and UPC without strict column spacing.
    """
    if not _looks_like_chinatown_invoice(lines_with_meta):
        return None

    errors: List[Dict] = []
    products: List[ProductCreate] = []
    products_by_id: Dict[str, Dict] = {}
    description_lines: List[str] = []
    table_started = False

    for idx, (raw_line, page_num) in enumerate(lines_with_meta):
        stripped = raw_line.strip()
        normalized = stripped.lower()

        if not table_started:
            if _looks_like_chinatown_header(normalized):
                table_started = True
                description_lines.clear()
            continue

        if not stripped:
            continue

        if _is_footer_line(normalized):
            description_lines.clear()
            continue

        if _looks_like_chinatown_header(normalized):
            # New page header inside the table; reset description buffer
            description_lines.clear()
            continue

        if _looks_like_chinatown_summary_line(stripped):
            parsed = _parse_chinatown_summary_line(stripped)
            if not parsed:
                errors.append({
                    "row": f"page {page_num}, line {idx + 1}",
                    "error": f"Could not parse summary line: {raw_line}"
                })
                description_lines.clear()
                continue

            price_value, product_id, pack_text = parsed

            description = _normalize_whitespace(" ".join(description_lines))
            description_lines.clear()

            if not description:
                description = pack_text or f"Item {product_id}"

            name = description
            if pack_text and pack_text.lower() not in name.lower():
                name = f"{name} {pack_text}".strip()
            name = _normalize_whitespace(name)

            if not product_id or price_value is None:
                errors.append({
                    "row": f"page {page_num}, line {idx + 1}",
                    "error": f"Incomplete summary data for line: {raw_line}"
                })
                continue

            if product_id in products_by_id:
                errors.append({
                    "row": f"page {page_num}, line {idx + 1}",
                    "error": f"Duplicate item code '{product_id}' detected in PDF. Using first occurrence."
                })
                continue

            product = ProductCreate(
                product_id=product_id,
                name=name,
                brand=None,
                price=price_value,
                stock=0,
                order_qty=0,
                image_url=None,
                remarks=None,
            )
            products.append(product)
            products_by_id[product_id] = {
                "product_id": product_id,
                "name": name,
                "price": price_value,
                "source": f"page {page_num}, line {idx + 1}",
            }
        else:
            description_lines.append(stripped)

    if not products:
        errors.append({
            "row": 0,
            "error": "Detected Chinatown invoice layout but could not extract any products."
        })

    return products, errors


def _looks_like_chinatown_invoice(lines_with_meta: List[Tuple[str, int]]) -> bool:
    joined = " ".join(line.lower() for line, _ in lines_with_meta[:80])
    return "chinatown supermarket" in joined and "es houseware" in joined


def _looks_like_chinatown_header(normalized_line: str) -> bool:
    normalized_line = normalized_line.replace(" ", "")
    return (
        "description" in normalized_line
        and "priceeach" in normalized_line
        and "itemcode" in normalized_line
    ) or (
        "totaldescription" in normalized_line
        and "itemcode" in normalized_line
    )


def _is_footer_line(normalized_line: str) -> bool:
    return (
        normalized_line.startswith("page ")
        or any(marker in normalized_line for marker in PDF_FOOTER_MARKERS)
    )


def _looks_like_chinatown_summary_line(line: str) -> bool:
    """
    Check if a line looks like a Chinatown invoice summary line.
    Summary lines contain price/amount values and typically have either:
    - A long digit sequence (barcode or numeric product ID)
    - An alphanumeric product ID pattern (like GT-099, JK51029, HOOK)
    """
    has_decimal = bool(PDF_CURRENCY_PATTERN.search(line))
    if not has_decimal:
        return False
    
    # Check for long digit sequences (barcodes or numeric product IDs)
    has_long_digits = bool(re.search(r"\d{5,}", line))
    
    # Check for alphanumeric product ID patterns (letters followed by optional dash and digits)
    # Examples: GT-099, JE-3345, JK51029, JKK283, JL8166
    has_alphanumeric_id = bool(re.search(r"[A-Z]{1,4}[-]?\d{2,5}\b", line))
    
    # Special case: all-letter product IDs like "HOOK" at the end after amount
    # Pattern: decimal amount (1+ digits, dot, 2 digits) immediately followed by 2-10 uppercase letters
    # Examples: "0.00HOOK", "60.00HOOK"
    has_letter_only_id = bool(re.search(r"\d+\.\d{2}[A-Z]{2,10}\b", line))
    
    return has_long_digits or has_alphanumeric_id or has_letter_only_id


def _parse_chinatown_summary_line(line: str) -> Optional[Tuple[Decimal, Optional[str], Optional[str]]]:
    """
    Parse a Chinatown invoice summary line to extract price, product ID, and pack text.
    
    Summary line format typically:
    [pack_text] [qty] [price] [amount] [product_id] [barcode]
    
    Examples:
    - '10PCS/BX, 10 BX/CS20 10.00 200.00GT-099 859176000518'
    - '12 PCS / BX, 288 PCS / CS72 1.50 108.00JE-3345 4930428833453'
    - '6PCS/ BX, 60PCS/CS6 13.00 78.00JK51029 4945569510293'
    - 'S/S HOOK 挂钩 20 0.00 0.00HOOK'
    """
    decimal_matches = list(PDF_CURRENCY_PATTERN.finditer(line))
    if not decimal_matches:
        return None

    price_value = _to_decimal(decimal_matches[0].group())
    amount_match = decimal_matches[1] if len(decimal_matches) > 1 else None
    tail_start = amount_match.end() if amount_match else decimal_matches[0].end()
    tail_text = line[tail_start:].strip()

    product_id = None
    if tail_text:
        # First, try to extract alphanumeric product IDs
        # Pattern: 1-4 letters, optional dash, 2-5 digits (e.g., GT-099, JE-3345, JK51029, JKK283)
        alphanumeric_match = re.match(r"([A-Z]{1,4}[-]?\d{2,5})\b", tail_text)
        if alphanumeric_match:
            product_id = alphanumeric_match.group(1)
        else:
            # Check for letter-only product IDs (e.g., HOOK)
            # These appear right after the amount value with no space
            letter_only_match = re.match(r"([A-Z]{2,10})\b", tail_text)
            if letter_only_match:
                candidate = letter_only_match.group(1)
                # Make sure it's not a common word/unit
                common_words = {"PCS", "BOX", "BX", "CS", "SET", "BG", "BD", "PC"}
                if candidate not in common_words:
                    product_id = candidate
            
            # If no alphanumeric ID found, try numeric ID
            if not product_id:
                id_match = re.search(r"(\d{5,8})", tail_text)
                if id_match:
                    product_id = id_match.group(1)[-6:]
    
    # Check if product ID is stuck to the amount value (no space between)
    # e.g., "200.00GT-099" or "0.00HOOK"
    if not product_id and amount_match:
        amount_end_pos = amount_match.end()
        remaining = line[amount_end_pos:]
        # Check if alphanumeric ID starts immediately after amount
        stuck_alpha_match = re.match(r"([A-Z]{1,4}[-]?\d{2,5})\b", remaining)
        if stuck_alpha_match:
            product_id = stuck_alpha_match.group(1)
        else:
            stuck_letter_match = re.match(r"([A-Z]{2,10})\b", remaining)
            if stuck_letter_match:
                candidate = stuck_letter_match.group(1)
                common_words = {"PCS", "BOX", "BX", "CS", "SET", "BG", "BD", "PC"}
                if candidate not in common_words:
                    product_id = candidate

    # Fallback: look for 6-digit numeric product ID anywhere in the line
    if not product_id:
        fallback_match = re.search(r"(\d{6})", line)
        if fallback_match:
            product_id = fallback_match.group(1)

    pack_text = line[:decimal_matches[0].start()].rstrip()
    if pack_text:
        pack_text = re.sub(r"\s*\d+\s*$", "", pack_text).strip()
        pack_text = _prettify_pack_text(pack_text)

    return price_value, product_id, pack_text or None


def _normalize_whitespace(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def _prettify_pack_text(value: Optional[str]) -> Optional[str]:
    if not value:
        return value
    beautified = re.sub(r"(?<=\d)(?=[A-Za-z])", " ", value)
    beautified = re.sub(r"(?<=[A-Za-z])(?=\d)", " ", beautified)
    return _normalize_whitespace(beautified)


def _looks_like_header(line: str) -> bool:
    normalized = line.lower()
    return "item" in normalized and "description" in normalized and ("price each" in normalized or "unit price" in normalized)


def _build_column_positions(header_line: str) -> Dict[str, int]:
    header_lower = header_line.lower()
    positions: Dict[str, int] = {}
    for key, keywords in PDF_HEADER_KEYWORDS.items():
        for keyword in keywords:
            idx = header_lower.find(keyword)
            if idx != -1:
                positions[key] = idx
                break
    return positions


def _locate_pdf_header(lines_with_meta: List[Tuple[str, int]]) -> Tuple[int, Dict[str, int]]:
    for idx, (line, _) in enumerate(lines_with_meta):
        if _looks_like_header(line):
            positions = _build_column_positions(line)
            if all(positions.get(field) is not None for field in ["product_id", "description", "price_each"]):
                return idx, positions
    return -1, {}


def _column_slice(line: str, start: Optional[int], end: Optional[int]) -> str:
    if start is None:
        return ""
    if start < 0:
        start = 0
    if end is not None and end < start:
        end = start
    return line[start:end].strip() if end is not None else line[start:].strip()


def _parse_line_by_columns(line: str, positions: Dict[str, int]) -> Optional[Tuple[str, str, Decimal]]:
    if (
        positions.get("product_id") is None
        or positions.get("description") is None
        or positions.get("price_each") is None
    ):
        return None
    
    product_id = _column_slice(line, positions.get("product_id"), positions.get("description"))
    
    desc_end_candidates = [
        positions.get("case_pack"),
        positions.get("qty"),
        positions.get("price_each"),
    ]
    desc_end_candidates = [pos for pos in desc_end_candidates if pos is not None and pos > positions["description"]]
    desc_end = min(desc_end_candidates) if desc_end_candidates else positions["price_each"]
    name = _column_slice(line, positions.get("description"), desc_end)

    case_pack_text = ""
    case_pack_start = positions.get("case_pack")
    if case_pack_start is not None:
        case_pack_end_candidates = [
            positions.get("qty"),
            positions.get("price_each"),
        ]
        case_pack_end_candidates = [
            pos for pos in case_pack_end_candidates
            if pos is not None and pos > case_pack_start
        ]
        case_pack_end = min(case_pack_end_candidates) if case_pack_end_candidates else None
        case_pack_text = _column_slice(line, case_pack_start, case_pack_end)
    
    price_end_candidates = [
        positions.get("amount"),
    ]
    price_end_candidates = [pos for pos in price_end_candidates if pos is not None and pos > positions["price_each"]]
    price_end = min(price_end_candidates) if price_end_candidates else None
    price_text = _column_slice(line, positions.get("price_each"), price_end)
    
    if not price_text:
        decimal_candidates = _extract_currency_values(line)
        price_value = decimal_candidates[-2] if len(decimal_candidates) >= 2 else (decimal_candidates[-1] if decimal_candidates else None)
    else:
        price_value = _to_decimal(price_text)
        if price_value is None:
            decimal_candidates = _extract_currency_values(line)
            price_value = decimal_candidates[-2] if len(decimal_candidates) >= 2 else (decimal_candidates[-1] if decimal_candidates else None)
    
    full_name = name.strip()
    normalized_case_pack = case_pack_text.strip()
    if normalized_case_pack and normalized_case_pack.lower() not in full_name.lower():
        full_name = f"{full_name} {normalized_case_pack}".strip()

    if not product_id or not full_name or price_value is None:
        return None

    return product_id.strip(), full_name, price_value


def _parse_line_by_pattern(line: str) -> Optional[Tuple[str, str, Decimal]]:
    decimal_candidates = _extract_currency_values(line)
    if not decimal_candidates:
        return None
    
    price_value = decimal_candidates[-2] if len(decimal_candidates) >= 2 else decimal_candidates[-1]
    price_match = list(PDF_CURRENCY_PATTERN.finditer(line))
    if not price_match:
        return None
    
    selected_match = price_match[-2] if len(price_match) >= 2 else price_match[-1]
    prefix = line[:selected_match.start()].rstrip()
    if not prefix:
        return None
    
    parts = re.split(r"\\s{2,}", prefix, maxsplit=1)
    if len(parts) < 2:
        return None
    
    product_id = parts[0].strip()
    name = parts[1].strip()
    
    if not product_id or not name:
        return None
    
    return product_id, name, price_value


def _extract_currency_values(line: str) -> List[Decimal]:
    values: List[Decimal] = []
    for match in PDF_CURRENCY_PATTERN.findall(line):
        decimal_value = _to_decimal(match)
        if decimal_value is not None:
            values.append(decimal_value)
    return values


def _to_decimal(value: str) -> Optional[Decimal]:
    if value is None:
        return None
    cleaned = value.strip()
    if not cleaned:
        return None
    is_negative = cleaned.startswith("(") and cleaned.endswith(")")
    cleaned = cleaned.strip("()")
    cleaned = cleaned.replace("$", "").replace(",", "").replace(" ", "")
    cleaned = cleaned.replace("USD", "").strip()
    if not cleaned:
        return None
    if is_negative and not cleaned.startswith("-"):
        cleaned = f"-{cleaned}"
    cleaned = re.sub(r"[^0-9.\\-]", "", cleaned)
    try:
        return Decimal(cleaned)
    except (InvalidOperation, ValueError):
        return None

